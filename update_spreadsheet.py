import pandas as pd
from openpyxl import load_workbook
import subprocess
import time
import os

# File paths
SPREADSHEET_PATH = "RTW_commanders_database.xlsx"
CSV_PATH = "latest_RTW_commanders_database.csv"

def update_spreadsheet():
    try:
        # Load the existing spreadsheet
        wb = load_workbook(SPREADSHEET_PATH)
        sheet = wb.active

        # Load CSV data
        df = pd.read_csv(CSV_PATH)

        # **Clear old data before inserting new data**
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

        # **Write new data from CSV**
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # **Force Git to detect changes**
        sheet["A1"] = f"Last update: {time.strftime('%Y-%m-%d %H:%M:%S')}"

        # Save the updated spreadsheet
        wb.save(SPREADSHEET_PATH)
        print("Spreadsheet successfully updated.")

    except Exception as e:
        print(f"Error updating spreadsheet: {e}")

def git_pull():
    """Pull latest changes from GitHub to avoid conflicts."""
    subprocess.run(["git", "config", "--global", "user.email", "github-actions[bot]@users.noreply.github.com"])
    subprocess.run(["git", "config", "--global", "user.name", "GitHub Actions Bot"])
    subprocess.run(["git", "pull", "--rebase", "origin", "main"])  # Prevent conflicts

def git_push():
    """Stage, commit, and push changes to GitHub."""
    subprocess.run(["git", "add", SPREADSHEET_PATH, CSV_PATH])

    # **Check if there are actual changes**
    status_output = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True)
    if not status_output.stdout.strip():
        print("No new changes detected. Skipping push.")
        return  # **Exit if no new data was added**

    # **Commit and force push**
    subprocess.run(["git", "commit", "-m", "Auto-updated commander database and CSV"])
    subprocess.run(["git", "push", "origin", "main", "--force"])  # Force push in case of sync issues

if __name__ == "__main__":
    git_pull()  # Ensure latest data is present
    update_spreadsheet()  # Update the spreadsheet with new commander data
    git_push()  # Push the updated files to GitHub
